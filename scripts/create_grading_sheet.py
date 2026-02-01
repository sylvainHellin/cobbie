#!/usr/bin/env python3
"""
Create Grading Sheet for Human Evaluation

This script exports evaluation data from MLflow into a structured Excel grading sheet
for inter-rater reliability analysis. It supports evaluation by 3 judges:
LLM-as-a-judge (pre-filled), Human Judge 1, and Human Judge 2.

Usage:
    uv run scripts/create_grading_sheet.py --run-ids <run_id1> <run_id2> ...
    uv run scripts/create_grading_sheet.py --run-ids c0f5d69f17b3400093fa63204c70adc3
    uv run scripts/create_grading_sheet.py --run-ids abc123 --output outputs/eval/my_grading.xlsx
"""

import argparse
import re
import sqlite3
from datetime import datetime
from typing import Dict, List

import mlflow
import numpy as np
import pandas as pd
from mlflow import MlflowClient
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.worksheet.datavalidation import DataValidation

from src.config import DB_PATH, MLFLOW_URI

# Constants
REPORTS_DIR = "outputs/eval"
CATEGORY_NAMES = {
    1: "Direct Property",
    2: "Aggregation",
    3: "Computation",
    4: "Estimation/Unavailable",
}

# Krippendorff's alpha will be imported when needed
try:
    import krippendorff
    KRIPPENDORFF_AVAILABLE = True
except ImportError:
    KRIPPENDORFF_AVAILABLE = False


# ============================================================================
# Utility Functions
# ============================================================================

def sanitize_for_excel(text: str) -> str:
    """
    Sanitize text to remove characters that are illegal in Excel cells.

    Excel/openpyxl doesn't allow certain control characters (0x00-0x1F except tab, newline, carriage return).
    Also truncates very long text to Excel's limit.

    Args:
        text: Input text string

    Returns:
        Sanitized text safe for Excel
    """
    if not isinstance(text, str):
        return str(text) if text is not None else ""

    # Remove illegal XML characters (Excel uses XML internally)
    illegal_chars = re.compile(r'[\x00-\x08\x0B\x0C\x0E-\x1F]')
    text = illegal_chars.sub('', text)

    # Truncate if too long (Excel cell limit is 32,767 characters)
    if len(text) > 32767:
        text = text[:32760] + "...[truncated]"

    return text


# ============================================================================
# Data Fetching Functions
# ============================================================================

def fetch_nested_runs(client: MlflowClient, parent_run_id: str, experiment_id: str) -> List:
    """
    Fetch all nested runs for a given parent run.

    Args:
        client: MLflow client instance
        parent_run_id: ID of the parent evaluation run
        experiment_id: ID of the experiment

    Returns:
        List of nested run objects
    """
    nested_runs = client.search_runs(
        experiment_ids=[experiment_id],
        filter_string=f'tags.mlflow.parentRunId = "{parent_run_id}"',
        max_results=1000,
    )
    return nested_runs


def extract_evaluation_data(run, parent_model_name: str = "Unknown") -> Dict:
    """
    Extract evaluation data from a single nested run.

    Args:
        run: MLflow run object
        parent_model_name: Model name from parent run

    Returns:
        Dictionary with extracted evaluation data
    """
    params = run.data.params
    tags = run.data.tags

    # Get question ID from run name (e.g., "question_0_909" -> 909)
    run_name = tags.get("mlflow.runName", "")
    question_id = None
    if run_name.startswith("question_"):
        try:
            # Format is "question_{index}_{id}"
            parts = run_name.split("_")
            if len(parts) >= 3:
                question_id = int(parts[2])
            else:
                question_id = int(parts[1])
        except (IndexError, ValueError):
            question_id = params.get("question_id")
    else:
        question_id = params.get("question_id")

    # Extract evaluation criteria
    abstention_str = params.get("abstention", "False")
    abstention = abstention_str == "True" or abstention_str == "true"

    faithfulness = params.get("faithfulness", "Na")
    completeness = params.get("completeness", "Na")
    transparency = params.get("transparency", "Na")
    relevance = params.get("relevance", "Na")

    # Extract answer and ground truth
    answer = params.get("answer", "")
    justification = params.get("justification", "")

    # Use model name from parent run
    model_name = parent_model_name

    data = {
        "question_id": question_id,
        "answer": answer,
        "justification": justification,
        "abstention": abstention,
        "faithfulness": faithfulness,
        "completeness": completeness,
        "transparency": transparency,
        "relevance": relevance,
        "model_name": model_name,
    }

    return data


def fetch_question_data(question_ids: List[int]) -> Dict[int, Dict]:
    """
    Fetch question data from the database.

    Args:
        question_ids: List of question IDs to fetch

    Returns:
        Dictionary mapping question_id to question data
    """
    if not question_ids:
        return {}

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    # Build query with JOIN to get project and model names
    placeholders = ",".join("?" * len(question_ids))
    query = f"""
        SELECT
            ib.id,
            ib.question,
            ib.ground_truth,
            ib.category,
            im.project_name,
            im.model_name
        FROM ifc_bench ib
        LEFT JOIN ifcmodels im ON ib.ifc_id = im.id
        WHERE ib.id IN ({placeholders})
    """

    cursor.execute(query, question_ids)
    rows = cursor.fetchall()
    conn.close()

    # Build dictionary
    question_data = {}
    for row in rows:
        question_data[row[0]] = {
            "question": row[1],
            "ground_truth": row[2],
            "category": row[3],
            "project_name": row[4],
            "model_name": row[5],
        }

    return question_data


# ============================================================================
# Data Processing Functions
# ============================================================================

def build_grading_dataframe(runs_data: List[Dict], question_data: Dict[int, Dict]) -> pd.DataFrame:
    """
    Build a pandas DataFrame from run data and question data for grading.

    Args:
        runs_data: List of dictionaries with evaluation data
        question_data: Dictionary mapping question_id to question data

    Returns:
        Pandas DataFrame with all data needed for grading sheet
    """
    rows = []

    for run_data in runs_data:
        question_id = run_data["question_id"]
        if question_id is None:
            continue

        # Get question data from database
        q_data = question_data.get(int(question_id), {})

        # Combine all data (sanitize text fields for Excel)
        row = {
            # Metadata
            "question_id": question_id,
            "question": sanitize_for_excel(q_data.get("question", "N/A")),
            "ground_truth": sanitize_for_excel(q_data.get("ground_truth", "N/A")),
            "category": q_data.get("category", 0),
            "category_name": CATEGORY_NAMES.get(q_data.get("category", 0), "Unknown"),
            "project_name": sanitize_for_excel(q_data.get("project_name", "N/A")),
            "model_name": sanitize_for_excel(run_data.get("model_name", "N/A")),
            # LLM's answer and evaluation
            "cobbie_answer": sanitize_for_excel(run_data["answer"]),
            "llm_abstention": run_data["abstention"],
            "llm_faithfulness": run_data["faithfulness"],
            "llm_completeness": run_data["completeness"],
            "llm_transparency": run_data["transparency"],
            "llm_relevance": run_data["relevance"],
            "llm_justification": sanitize_for_excel(run_data["justification"]),
        }

        rows.append(row)

    df = pd.DataFrame(rows)
    return df


def derive_binary_classification(abstention: bool, faithfulness: str, completeness: str) -> str:
    """
    Derive binary classification from criteria.

    Args:
        abstention: Whether system abstained
        faithfulness: Faithfulness criterion (Yes/No/Na)
        completeness: Completeness criterion (Yes/No/Na)

    Returns:
        Binary classification: "abstained", "correct", or "wrong"
    """
    if abstention:
        return "abstained"
    elif faithfulness == "Yes" and completeness == "Yes":
        return "correct"
    else:
        return "wrong"


# ============================================================================
# Encoding Functions for Krippendorff's Alpha
# ============================================================================

def encode_abstention(value) -> float:
    """Encode abstention boolean for Krippendorff's alpha."""
    if pd.isna(value) or value is None:
        return np.nan
    return 1.0 if value else 0.0


def encode_criterion(value: str) -> float:
    """Encode Yes/No/Na criterion for Krippendorff's alpha (ordinal)."""
    if pd.isna(value) or value is None:
        return np.nan
    mapping = {"Yes": 2.0, "Na": 1.0, "No": 0.0}
    return mapping.get(value, np.nan)


def encode_binary(value: str) -> float:
    """Encode binary classification for Krippendorff's alpha (nominal)."""
    if pd.isna(value) or value is None:
        return np.nan
    mapping = {"correct": 2.0, "abstained": 1.0, "wrong": 0.0}
    return mapping.get(value, np.nan)


# ============================================================================
# Krippendorff's Alpha Calculation
# ============================================================================

def calculate_criterion_alpha(df: pd.DataFrame, criterion: str, judges: List[str] | None = None) -> Dict[str, float]:
    """
    Calculate Krippendorff's alpha for a criterion across judges.

    Args:
        df: DataFrame with judge evaluations
        criterion: Name of criterion (abstention, faithfulness, completeness, etc.)
        judges: List of judge prefixes (default: ['llm', 'human1', 'human2'])

    Returns:
        Dictionary with combined and pairwise alpha values
    """
    if not KRIPPENDORFF_AVAILABLE:
        return {
            'combined': np.nan,
            'llm_human1': np.nan,
            'llm_human2': np.nan,
            'human1_human2': np.nan,
        }

    if judges is None:
        judges = ['llm', 'human1', 'human2']

    # Select encoding function
    if criterion == 'abstention':
        encode_func = encode_abstention
        level = 'nominal'
    elif criterion == 'binary':
        encode_func = encode_binary
        level = 'nominal'
    else:
        encode_func = encode_criterion
        level = 'ordinal'

    # Build reliability data matrix (judges × questions)
    reliability_data = []
    for judge in judges:
        col_name = f"{judge}_{criterion}"
        if col_name in df.columns:
            values = df[col_name].apply(encode_func)
            reliability_data.append(values.to_numpy())
        else:
            # Judge hasn't filled this column yet
            reliability_data.append(np.full(len(df), np.nan))

    # Check if we have enough data
    reliability_array = np.array(reliability_data)
    valid_cols = ~np.all(np.isnan(reliability_array), axis=0)
    if np.sum(valid_cols) < 2:
        # Not enough data
        return {
            'combined': np.nan,
            'llm_human1': np.nan,
            'llm_human2': np.nan,
            'human1_human2': np.nan,
        }

    try:
        # Combined alpha (all judges)
        combined_alpha = krippendorff.alpha(
            reliability_data,
            level_of_measurement=level
        )

        # Pairwise alphas
        llm_h1_alpha = krippendorff.alpha(
            [reliability_data[0], reliability_data[1]],
            level_of_measurement=level
        )

        llm_h2_alpha = krippendorff.alpha(
            [reliability_data[0], reliability_data[2]],
            level_of_measurement=level
        )

        h1_h2_alpha = krippendorff.alpha(
            [reliability_data[1], reliability_data[2]],
            level_of_measurement=level
        )

        return {
            'combined': combined_alpha if not np.isnan(combined_alpha) else np.nan,
            'llm_human1': llm_h1_alpha if not np.isnan(llm_h1_alpha) else np.nan,
            'llm_human2': llm_h2_alpha if not np.isnan(llm_h2_alpha) else np.nan,
            'human1_human2': h1_h2_alpha if not np.isnan(h1_h2_alpha) else np.nan,
        }
    except Exception as e:
        print(f"  Warning: Error calculating alpha for {criterion}: {e}")
        return {
            'combined': np.nan,
            'llm_human1': np.nan,
            'llm_human2': np.nan,
            'human1_human2': np.nan,
        }


def calculate_all_alphas(df: pd.DataFrame) -> pd.DataFrame:
    """
    Calculate Krippendorff's alpha for all criteria.

    Args:
        df: DataFrame with judge evaluations (must have binary columns added)

    Returns:
        DataFrame with alpha values for each criterion
    """
    criteria = ['abstention', 'faithfulness', 'completeness', 'transparency', 'relevance', 'binary']

    alpha_data = []
    for criterion in criteria:
        alphas = calculate_criterion_alpha(df, criterion)
        alpha_data.append({
            'Criterion': criterion.capitalize(),
            'Combined α (3 judges)': alphas['combined'],
            'LLM-Human1 α': alphas['llm_human1'],
            'LLM-Human2 α': alphas['llm_human2'],
            'Human1-Human2 α': alphas['human1_human2'],
        })

    return pd.DataFrame(alpha_data)


# ============================================================================
# Agreement Statistics Functions
# ============================================================================

def calculate_percentage_agreement(df: pd.DataFrame, criterion: str) -> Dict:
    """Calculate percentage agreement between judge pairs."""
    agreements = {}

    pairs = [
        ('llm', 'human1'),
        ('llm', 'human2'),
        ('human1', 'human2'),
    ]

    for judge1, judge2 in pairs:
        col1 = f"{judge1}_{criterion}"
        col2 = f"{judge2}_{criterion}"

        if col1 in df.columns and col2 in df.columns:
            # Count where both judges have values and they agree
            both_valid = df[[col1, col2]].notna().all(axis=1)
            total = int(both_valid.sum())
            if total > 0:
                agree = int((df.loc[both_valid, col1] == df.loc[both_valid, col2]).sum())
                pct = agree / total if total > 0 else 0
                agreements[f"{judge1}_{judge2}"] = {
                    'agree': agree,
                    'total': total,
                    'percentage': pct
                }
            else:
                agreements[f"{judge1}_{judge2}"] = {
                    'agree': 0,
                    'total': 0,
                    'percentage': 0
                }
        else:
            agreements[f"{judge1}_{judge2}"] = {
                'agree': 0,
                'total': 0,
                'percentage': 0
            }

    # Overall 3-way agreement
    if f"llm_{criterion}" in df.columns and f"human1_{criterion}" in df.columns and f"human2_{criterion}" in df.columns:
        all_valid = df[[f"llm_{criterion}", f"human1_{criterion}", f"human2_{criterion}"]].notna().all(axis=1)
        total_all = int(all_valid.sum())
        if total_all > 0:
            all_agree = int((
                (df.loc[all_valid, f"llm_{criterion}"] == df.loc[all_valid, f"human1_{criterion}"]) &
                (df.loc[all_valid, f"llm_{criterion}"] == df.loc[all_valid, f"human2_{criterion}"])
            ).sum())
            agreements['all_judges'] = {
                'agree': all_agree,
                'total': total_all,
                'percentage': all_agree / total_all if total_all > 0 else 0
            }
        else:
            agreements['all_judges'] = {'agree': 0, 'total': 0, 'percentage': 0}
    else:
        agreements['all_judges'] = {'agree': 0, 'total': 0, 'percentage': 0}

    return agreements


# ============================================================================
# Excel Creation Functions
# ============================================================================

def create_instructions_sheet(wb: Workbook) -> None:
    """Create the Instructions sheet with evaluation guidelines."""
    ws = wb.create_sheet("Instructions", 0)

    instructions = [
        ["Grading Sheet Instructions", ""],
        ["", ""],
        ["1. Overview", ""],
        ["This grading sheet is designed for human evaluation of Cobbie's answers to BIM-related questions.", ""],
        ["You will evaluate answers using the same 5 criteria that the LLM judge uses:", ""],
        ["  - Abstention (Boolean): Did the system decline to answer?", ""],
        ["  - Faithfulness (Yes/No/Na): Are all claims grounded in valid sources?", ""],
        ["  - Completeness (Yes/No/Na): Are all relevant facts included?", ""],
        ["  - Transparency (Yes/No/Na): Are sources/methods explicitly disclosed?", ""],
        ["  - Relevance (Yes/No/Na): Does the answer directly address the question?", ""],
        ["", ""],
        ["2. How to Use This Sheet", ""],
        ["  1. Go to your corresponding evaluation sheet", ""],
        ["  2. Fill in your evaluation in either 'Human Judge 1' or 'Human Judge 2' sheet", ""],
        ["  3. Use the dropdown menus for Faithfulness, Completeness, Transparency, Relevance", ""],
        ["  4. Check/uncheck the Abstention box as appropriate", ""],
        ["  5. Optionally provide justification for your ratings", ""],
        ["  6. The Binary Classification column will auto-calculate based on your ratings", ""],
        ["", ""],
        ["3. Evaluation Criteria Details", ""],
        ["", ""],
        ["Abstention:", ""],
        ["  TRUE: System explicitly declined to answer (e.g., 'I cannot determine...', 'Insufficient information...')", ""],
        ["  FALSE: System provided an answer", ""],
        ["", ""],
        ["Faithfulness:", ""],
        ["  Yes: All claims are grounded in valid sources for this category", ""],
        ["  No: Some claims are not properly grounded", ""],
        ["  Na: Only if Abstention is TRUE", ""],
        ["  Category-specific rules:", ""],
        ["    - Cat 1: Only BIM element properties", ""],
        ["    - Cat 2: Simple computations (count, sum, average)", ""],
        ["    - Cat 3: Complex geometric computations", ""],
        ["    - Cat 4: BIM data + EXPLICITLY STATED assumptions", ""],
        ["", ""],
        ["Completeness:", ""],
        ["  Yes: All relevant facts are included", ""],
        ["  No: Some relevant facts are missing", ""],
        ["  Na: If Abstention is TRUE OR question is open-ended with no objective completeness standard", ""],
        ["", ""],
        ["Transparency:", ""],
        ["  Yes: Sources/methods are explicitly disclosed for each claim", ""],
        ["  No: Some sources/methods are not disclosed", ""],
        ["  Na: Only if Abstention is TRUE", ""],
        ["", ""],
        ["Relevance:", ""],
        ["  Yes: Answer directly addresses the question asked", ""],
        ["  No: Answer does not address the question", ""],
        ["  Na: Only if Abstention is TRUE", ""],
        ["", ""],
        ["4. Binary Classification", ""],
        ["The binary classification is automatically derived from your ratings:", ""],
        ["  - If Abstention = TRUE → 'abstained'", ""],
        ["  - If Faithfulness = Yes AND Completeness = Yes → 'correct'", ""],
        ["  - Otherwise → 'wrong'", ""],
        ["", ""],
        ["5. Question Categories", ""],
        ["  Cat 1 - Direct Property: Lookup of BIM element properties", ""],
        ["  Cat 2 - Aggregation: Counting, summing, averaging", ""],
        ["  Cat 3 - Computation: Complex geometric calculations", ""],
        ["  Cat 4 - Estimation/Unavailable: Requires assumptions or data not in model", ""],
        ["", ""],
        ["6. Numerical Tolerance", ""],
        ["For numerical answers:", ""],
        ["  - Continuous values: ±2% tolerance", ""],
        ["  - Discrete values (counts): Must be exact", ""],
    ]

    for row_idx, (text, _) in enumerate(instructions, start=1):
        ws.cell(row=row_idx, column=1, value=text)

    # Format
    ws.column_dimensions['A'].width = 120
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.font = Font(bold=True, size=14)


def create_judge_sheet(wb: Workbook, sheet_name: str, df: pd.DataFrame, judge_prefix: str, editable: bool) -> None:
    """
    Create a judge sheet (LLM or Human).

    Args:
        wb: Workbook object
        sheet_name: Name for the sheet
        df: DataFrame with grading data
        judge_prefix: Prefix for judge columns ('llm', 'human1', 'human2')
        editable: If True, make criteria columns editable for humans
    """
    ws = wb.create_sheet(sheet_name)

    # Define columns
    headers = [
        "Question ID",
        "Question",
        "Cobbie's Answer",
        "Ground Truth",
        "Category",
        "Category Name",
        "Project",
        "Model",
        "Abstention",
        "Faithfulness",
        "Completeness",
        "Transparency",
        "Relevance",
        "Justification",
        "Binary Classification",
    ]

    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Set column widths
    ws.column_dimensions['A'].width = 12  # Question ID
    ws.column_dimensions['B'].width = 50  # Question
    ws.column_dimensions['C'].width = 60  # Answer
    ws.column_dimensions['D'].width = 50  # Ground Truth
    ws.column_dimensions['E'].width = 10  # Category
    ws.column_dimensions['F'].width = 20  # Category Name
    ws.column_dimensions['G'].width = 25  # Project
    ws.column_dimensions['H'].width = 25  # Model
    ws.column_dimensions['I'].width = 12  # Abstention
    ws.column_dimensions['J'].width = 12  # Faithfulness
    ws.column_dimensions['K'].width = 12  # Completeness
    ws.column_dimensions['L'].width = 12  # Transparency
    ws.column_dimensions['M'].width = 12  # Relevance
    ws.column_dimensions['N'].width = 40  # Justification
    ws.column_dimensions['O'].width = 15  # Binary

    # Write data rows
    for row_idx, row_data in enumerate(df.itertuples(index=False), start=2):
        # Metadata columns (A-H)
        ws.cell(row=row_idx, column=1, value=row_data.question_id)
        ws.cell(row=row_idx, column=2, value=row_data.question).alignment = Alignment(wrap_text=True)
        ws.cell(row=row_idx, column=3, value=row_data.cobbie_answer).alignment = Alignment(wrap_text=True)
        ws.cell(row=row_idx, column=4, value=row_data.ground_truth).alignment = Alignment(wrap_text=True)
        ws.cell(row=row_idx, column=5, value=row_data.category)
        ws.cell(row=row_idx, column=6, value=row_data.category_name)
        ws.cell(row=row_idx, column=7, value=row_data.project_name)
        ws.cell(row=row_idx, column=8, value=row_data.model_name)

        # Criteria columns (I-M) - fill with LLM data or leave empty
        if judge_prefix == 'llm':
            # Pre-fill with LLM's grades
            ws.cell(row=row_idx, column=9, value=row_data.llm_abstention)
            ws.cell(row=row_idx, column=10, value=row_data.llm_faithfulness)
            ws.cell(row=row_idx, column=11, value=row_data.llm_completeness)
            ws.cell(row=row_idx, column=12, value=row_data.llm_transparency)
            ws.cell(row=row_idx, column=13, value=row_data.llm_relevance)
            ws.cell(row=row_idx, column=14, value=row_data.llm_justification).alignment = Alignment(wrap_text=True)

            # Binary classification formula
            binary_formula = f'=IF(I{row_idx},"abstained",IF(AND(J{row_idx}="Yes",K{row_idx}="Yes"),"correct","wrong"))'
            ws.cell(row=row_idx, column=15, value=binary_formula)
        else:
            # Leave empty for human judges to fill in
            # All criteria cells are left blank (no default values)
            ws.cell(row=row_idx, column=9, value=None)  # Abstention
            ws.cell(row=row_idx, column=10, value=None)  # Faithfulness
            ws.cell(row=row_idx, column=11, value=None)  # Completeness
            ws.cell(row=row_idx, column=12, value=None)  # Transparency
            ws.cell(row=row_idx, column=13, value=None)  # Relevance
            ws.cell(row=row_idx, column=14, value="").alignment = Alignment(wrap_text=True)  # Justification

            # Binary classification formula (will show errors until human fills criteria)
            binary_formula = f'=IF(I{row_idx},"abstained",IF(AND(J{row_idx}="Yes",K{row_idx}="Yes"),"correct","wrong"))'
            ws.cell(row=row_idx, column=15, value=binary_formula)

        # Center align criteria columns
        for col in range(9, 14):
            ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row_idx, column=15).alignment = Alignment(horizontal="center", vertical="center")

    # Add data validation for human sheets
    if editable:
        # Boolean validation for Abstention (column I)
        dv_bool = DataValidation(
            type="list",
            formula1='"TRUE,FALSE"',
            allow_blank=False,
            showErrorMessage=True,
            error="Please select TRUE or FALSE",
            errorTitle="Invalid Entry"
        )
        ws.add_data_validation(dv_bool)
        dv_bool.add(f'I2:I{len(df)+1}')

        # Yes/No/Na validation for criteria (columns J-M)
        dv_criteria = DataValidation(
            type="list",
            formula1='"Yes,No,Na"',
            allow_blank=False,
            showErrorMessage=True,
            error="Please select Yes, No, or Na",
            errorTitle="Invalid Entry"
        )
        ws.add_data_validation(dv_criteria)
        dv_criteria.add(f'J2:M{len(df)+1}')

        # Unlock editable cells (I-N: criteria and justification)
        for row in range(2, len(df) + 2):
            for col in range(9, 15):  # I through N
                ws.cell(row=row, column=col).protection = Protection(locked=False)

        # Protect the sheet (allows editing unlocked cells only)
        ws.protection.sheet = True
        ws.protection.selectLockedCells = True
        ws.protection.selectUnlockedCells = True


def create_alpha_summary_sheet(wb: Workbook, alpha_df: pd.DataFrame, df: pd.DataFrame) -> None:
    """Create the Krippendorff's Alpha Summary sheet."""
    ws = wb.create_sheet("Krippendorff's Alpha Summary")

    # Check completion status
    total_questions = len(df)
    llm_complete = total_questions  # LLM always complete

    human1_complete = 0
    human2_complete = 0

    # Count how many questions each human has completed (all criteria filled)
    for _, row in df.iterrows():
        # For human1, check if they have any non-default values
        # Since we don't have human columns in initial df, show 0
        pass

    # Write completion status
    ws.cell(row=1, column=1, value="Completion Status")
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws.cell(row=2, column=1, value=f"LLM Judge: {llm_complete}/{total_questions} (100%)")
    ws.cell(row=3, column=1, value=f"Human Judge 1: {human1_complete}/{total_questions} (0% - awaiting input)")
    ws.cell(row=4, column=1, value=f"Human Judge 2: {human2_complete}/{total_questions} (0% - awaiting input)")

    # Write interpretation guide
    ws.cell(row=6, column=1, value="Interpretation Guide:")
    ws.cell(row=6, column=1).font = Font(bold=True)
    ws.cell(row=7, column=1, value="α > 0.800: Excellent reliability")
    ws.cell(row=8, column=1, value="α 0.667-0.800: Good reliability")
    ws.cell(row=9, column=1, value="α 0.500-0.667: Moderate reliability")
    ws.cell(row=10, column=1, value="α < 0.500: Poor reliability")

    # Write alpha values table
    start_row = 12
    ws.cell(row=start_row, column=1, value="Criterion")
    ws.cell(row=start_row, column=2, value="Combined α (3 judges)")
    ws.cell(row=start_row, column=3, value="LLM-Human1 α")
    ws.cell(row=start_row, column=4, value="LLM-Human2 α")
    ws.cell(row=start_row, column=5, value="Human1-Human2 α")

    for col in range(1, 6):
        ws.cell(row=start_row, column=col).font = Font(bold=True)
        ws.cell(row=start_row, column=col).fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    for idx, row_data in enumerate(alpha_df.itertuples(index=False), start=start_row+1):
        ws.cell(row=idx, column=1, value=row_data.Criterion)

        # Format alpha values
        for col_idx, val in enumerate([row_data._1, row_data._2, row_data._3, row_data._4], start=2):
            if pd.isna(val):
                ws.cell(row=idx, column=col_idx, value="N/A")
            else:
                ws.cell(row=idx, column=col_idx, value=f"{val:.3f}")
                # Color coding
                if val >= 0.667:
                    ws.cell(row=idx, column=col_idx).fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                elif val >= 0.500:
                    ws.cell(row=idx, column=col_idx).fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
                else:
                    ws.cell(row=idx, column=col_idx).fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")

    # Set column widths
    ws.column_dimensions['A'].width = 20
    for col in ['B', 'C', 'D', 'E']:
        ws.column_dimensions[col].width = 18


def create_binary_comparison_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    """Create the Binary Classification Comparison sheet."""
    ws = wb.create_sheet("Binary Classification")

    # Headers
    headers = ["Question ID", "Question (truncated)", "LLM Binary", "Human1 Binary", "Human2 Binary", "Agreement"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # Data rows
    for row_idx, row_data in enumerate(df.itertuples(index=False), start=2):
        # Question ID
        ws.cell(row=row_idx, column=1, value=row_data.question_id)

        # Truncated question (first 50 chars)
        question_truncated = row_data.question[:50] + "..." if len(row_data.question) > 50 else row_data.question
        ws.cell(row=row_idx, column=2, value=question_truncated)

        # Binary classifications (formulas referencing judge sheets)
        # For now, just show LLM's binary
        llm_binary = derive_binary_classification(
            row_data.llm_abstention,
            row_data.llm_faithfulness,
            row_data.llm_completeness
        )
        ws.cell(row=row_idx, column=3, value=llm_binary)
        ws.cell(row=row_idx, column=4, value="N/A")  # Human1 - to be filled
        ws.cell(row=row_idx, column=5, value="N/A")  # Human2 - to be filled
        ws.cell(row=row_idx, column=6, value="Awaiting human input")

    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 20

    # Summary statistics
    summary_row = len(df) + 3
    ws.cell(row=summary_row, column=1, value="Summary:")
    ws.cell(row=summary_row, column=1).font = Font(bold=True)
    ws.cell(row=summary_row+1, column=1, value=f"Total Questions: {len(df)}")
    ws.cell(row=summary_row+2, column=1, value="All Agree: Awaiting human input")
    ws.cell(row=summary_row+3, column=1, value="Partial Agreement: Awaiting human input")
    ws.cell(row=summary_row+4, column=1, value="No Agreement: Awaiting human input")


def create_agreement_stats_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    """Create the Agreement Statistics sheet."""
    ws = wb.create_sheet("Agreement Statistics")

    ws.cell(row=1, column=1, value="Detailed agreement statistics will be calculated after human evaluation is complete.")
    ws.cell(row=1, column=1).font = Font(bold=True)

    ws.cell(row=3, column=1, value="This sheet will contain:")
    ws.cell(row=4, column=1, value="  - Percentage agreement per criterion")
    ws.cell(row=5, column=1, value="  - Confusion matrices (pairwise)")
    ws.cell(row=6, column=1, value="  - Disagreement patterns")
    ws.cell(row=7, column=1, value="  - Category-level agreement breakdown")


# ============================================================================
# Main Function
# ============================================================================

def main():
    """Main function."""
    parser = argparse.ArgumentParser(
        description="Create grading sheet from MLflow evaluation runs",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Create grading sheet from single run
  uv run scripts/create_grading_sheet.py --run-ids c0f5d69f17b3400093fa63204c70adc3

  # Create from multiple runs
  uv run scripts/create_grading_sheet.py --run-ids run1 run2 run3

  # Custom output path
  uv run scripts/create_grading_sheet.py --run-ids abc123 --output outputs/eval/my_grading.xlsx
        """,
    )

    parser.add_argument(
        "--run-ids",
        nargs="+",
        required=True,
        help="MLflow evaluation run IDs to include (space-separated)",
    )

    parser.add_argument(
        "--output",
        help="Custom output path (default: outputs/eval/grading_sheet_YYYY-MM-DD_HH-MM-SS.xlsx)",
    )

    args = parser.parse_args()

    # Setup output path
    if args.output:
        output_path = args.output
    else:
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        output_path = f"{REPORTS_DIR}/grading_sheet_{timestamp}.xlsx"

    print("=" * 80)
    print("Creating Grading Sheet for Human Evaluation")
    print("=" * 80)
    print(f"\nProcessing {len(args.run_ids)} run(s)...")

    # Setup MLflow
    mlflow.set_tracking_uri(MLFLOW_URI)
    client = MlflowClient()

    # Collect all run data
    all_runs_data = []

    for run_id in args.run_ids:
        print(f"\nFetching data from run: {run_id}")

        try:
            # Get main run info
            main_run = client.get_run(run_id)
            experiment_id = main_run.info.experiment_id
            run_name = main_run.data.tags.get("mlflow.runName", "Unknown")

            # Extract model name from parent run
            parent_model_name = main_run.data.params.get("model_name",
                                main_run.data.tags.get("model_name", "Unknown"))

            print(f"  Run Name: {run_name}")
            print(f"  Experiment ID: {experiment_id}")
            print(f"  Model: {parent_model_name}")

            # Fetch nested runs
            nested_runs = fetch_nested_runs(client, run_id, experiment_id)
            print(f"  Found {len(nested_runs)} nested question runs")

            # Extract evaluation data from each nested run
            for nested_run in nested_runs:
                run_data = extract_evaluation_data(nested_run, parent_model_name)
                all_runs_data.append(run_data)

        except Exception as e:
            print(f"  Error processing run {run_id}: {e}")
            continue

    if not all_runs_data:
        print("\nError: No evaluation data found in the provided runs.")
        print("Please check that the run IDs are correct and contain nested question runs.")
        return

    print(f"\nTotal questions collected: {len(all_runs_data)}")

    # Fetch question data from database
    question_ids = [r["question_id"] for r in all_runs_data if r["question_id"] is not None]
    unique_question_ids = list(set(question_ids))
    print(f"Fetching metadata for {len(unique_question_ids)} unique questions from database...")
    question_data = fetch_question_data(unique_question_ids)

    # Build grading DataFrame
    print("Building grading DataFrame...")
    df = build_grading_dataframe(all_runs_data, question_data)

    # Category breakdown
    category_counts = df['category'].value_counts().to_dict()
    print("\nQuestions by category:")
    for cat, count in sorted(category_counts.items()):
        print(f"  {cat} - {CATEGORY_NAMES.get(cat, 'Unknown')}: {count}")

    # Add binary classification columns for LLM
    df['llm_binary'] = df.apply(
        lambda row: derive_binary_classification(
            row['llm_abstention'],
            row['llm_faithfulness'],
            row['llm_completeness']
        ),
        axis=1
    )

    # Add placeholder columns for humans (will be empty initially)
    for judge in ['human1', 'human2']:
        df[f'{judge}_abstention'] = np.nan
        df[f'{judge}_faithfulness'] = np.nan
        df[f'{judge}_completeness'] = np.nan
        df[f'{judge}_transparency'] = np.nan
        df[f'{judge}_relevance'] = np.nan
        df[f'{judge}_justification'] = ""
        df[f'{judge}_binary'] = np.nan

    # Calculate alphas (will be NaN for humans initially)
    print("Calculating Krippendorff's alpha...")
    alpha_df = calculate_all_alphas(df)

    # Create Excel workbook
    print("\nCreating Excel workbook...")
    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)  # Remove default sheet

    # Create all sheets
    print("  ✓ Instructions sheet")
    create_instructions_sheet(wb)

    print(f"  ✓ LLM Judge sheet ({len(df)} questions)")
    create_judge_sheet(wb, "LLM Judge", df, "llm", editable=False)

    print("  ✓ Human Judge 1 sheet (empty, ready for input)")
    create_judge_sheet(wb, "Human Judge 1", df, "human1", editable=True)

    print("  ✓ Human Judge 2 sheet (empty, ready for input)")
    create_judge_sheet(wb, "Human Judge 2", df, "human2", editable=True)

    print("  ✓ Krippendorff's Alpha Summary (awaiting human input)")
    create_alpha_summary_sheet(wb, alpha_df, df)

    print("  ✓ Binary Classification Comparison")
    create_binary_comparison_sheet(wb, df)

    print("  ✓ Agreement Statistics (placeholder)")
    create_agreement_stats_sheet(wb, df)

    # Save workbook
    print(f"\nSaving to: {output_path}")
    wb.save(output_path)

    # Summary
    print("\n" + "=" * 80)
    print("Summary:")
    print("=" * 80)
    print(f"  Total Questions: {len(df)}")
    print(f"  LLM Evaluations: {len(df)} (100%)")
    print("  Human 1 Evaluations: 0 (0%)")
    print("  Human 2 Evaluations: 0 (0%)")
    print("\n  Categories:")
    for cat, count in sorted(category_counts.items()):
        print(f"    {cat} - {CATEGORY_NAMES.get(cat, 'Unknown')}: {count}")

    print("\n" + "=" * 80)
    print("Next steps:")
    print("=" * 80)
    print("  1. Open the Excel file")
    print("  2. Fill in 'Human Judge 1' and/or 'Human Judge 2' sheets")
    print("  3. Use the dropdown menus for evaluation criteria")
    print("  4. Save the file")
    print("  5. Future: Re-run with --read-from to calculate inter-rater reliability")
    print("\n✅ Grading sheet created successfully!")


if __name__ == "__main__":
    main()
