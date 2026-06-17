"""One-off generator for the MiniMax cost-probe workbook.

Sources the dev-midi (40 questions/cell) aggregates from the existing report
.agents/research/2026-06-17-minimax-cost-probe.md. Derived, extrapolation, and
pricing cells are real worksheet formulas so the user can edit inputs
(testset_size, provider prices) and recompute. Run with:

    uv run --with openpyxl python scripts/_gen_cost_xlsx.py
"""

from __future__ import annotations

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

OUT = ".agents/research/2026-06-17-minimax-cost-probe.xlsx"

BOLD = Font(bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")

# dev-midi per-cell aggregates from the report (40 questions/cell, MiniMax-M3
# via the Anthropic endpoint, stratified by category with 10 per category).
# Order: cell_id, paradigm, tools, n, mean_input, mean_cached, mean_output,
# mean_latency_s, mean_tool_calls, mean_iterations.
CELLS = [
    ("minimax-anthropic-MiniMax-M3__agentic__none", "agentic", "none",
     40, 135100.8, 117687.2, 2999.7, 99.38, 8.25, 8.25),
    ("minimax-anthropic-MiniMax-M3__agentic__tools", "agentic", "tools",
     40, 169234.6, 148941.1, 3698.4, 106.81, 9.90, 9.90),
    ("minimax-anthropic-MiniMax-M3__static__none", "static", "none",
     40, 13195.9, 7105.4, 896.2, 21.77, 0.97, 0.97),
    ("minimax-anthropic-MiniMax-M3__static__tools", "static", "tools",
     40, 15084.9, 8409.8, 783.6, 21.09, 1.00, 1.00),
]


def _style_header(ws, row=1):
    for cell in ws[row]:
        cell.font = BOLD
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def _widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_per_cell(wb):
    ws = wb.active
    ws.title = "per_cell"
    headers = [
        "cell_id", "paradigm", "tools", "n_questions",
        "mean_input_tokens", "mean_cached_input_tokens", "mean_output_tokens",
        "mean_uncached_input_tokens", "mean_latency_s",
        "mean_tool_calls", "mean_iterations",
    ]
    ws.append(headers)
    for c in CELLS:
        cell_id, paradigm, tools, n, mi, mc, mo, lat, tc, it = c
        r = ws.max_row + 1
        ws.append([cell_id, paradigm, tools, n, mi, mc, mo, None, lat, tc, it])
        # mean_uncached_input = mean_input - mean_cached (column H = E - F)
        ws.cell(row=r, column=8).value = f"=E{r}-F{r}"
    _style_header(ws)
    _widths(ws, [46, 10, 8, 12, 18, 22, 18, 24, 14, 16, 16])
    return ws


def build_extrapolation(wb):
    ws = wb.create_sheet("extrapolation")
    # Single editable testset_size input cell, referenced by every row.
    ws["A1"] = "testset_size (edit this)"
    ws["A1"].font = BOLD
    ws["B1"] = 514
    ws["B1"].font = BOLD

    header_row = 3
    headers = [
        "cell_id", "testset_size", "mean_input", "mean_cached", "mean_output",
        "est_total_input", "est_total_cached", "est_total_output",
        "est_total_uncached_input",
    ]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col, value=h).font = BOLD

    first = header_row + 1  # row 4
    for i, c in enumerate(CELLS):
        r = first + i
        per_cell_row = 2 + i  # per_cell data rows start at 2
        ws.cell(row=r, column=1, value=c[0])
        ws.cell(row=r, column=2, value="=$B$1")
        ws.cell(row=r, column=3, value=f"=per_cell!E{per_cell_row}")
        ws.cell(row=r, column=4, value=f"=per_cell!F{per_cell_row}")
        ws.cell(row=r, column=5, value=f"=per_cell!G{per_cell_row}")
        ws.cell(row=r, column=6, value=f"=C{r}*B{r}")  # est_total_input
        ws.cell(row=r, column=7, value=f"=D{r}*B{r}")  # est_total_cached
        ws.cell(row=r, column=8, value=f"=E{r}*B{r}")  # est_total_output
        ws.cell(row=r, column=9, value=f"=F{r}-G{r}")  # est_total_uncached

    last = first + len(CELLS) - 1  # row 7
    total = last + 1  # row 8
    ws.cell(row=total, column=1, value="TOTAL").font = BOLD
    for col in (6, 7, 8, 9):
        L = get_column_letter(col)
        cell = ws.cell(row=total, column=col, value=f"=SUM({L}{first}:{L}{last})")
        cell.font = BOLD

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    _widths(ws, [46, 14, 14, 14, 14, 18, 18, 18, 24])
    return ws, first, last, total


def build_pricing(wb, ext_first, ext_last, ext_total):
    ws = wb.create_sheet("pricing_calc")

    # Section 1: active price set (editable, blank for the user to fill).
    ws["A1"] = "Active price set (USD per 1M tokens, edit these)"
    ws["A1"].font = BOLD
    ws["A2"] = "price_per_1M_input"
    ws["A3"] = "price_per_1M_cached_input"
    ws["A4"] = "price_per_1M_output"
    for r in (2, 3, 4):
        ws.cell(row=r, column=2).value = None  # blank input

    # Section 2: per-cell cost using the active price set, referencing the
    # extrapolation totals. cost = uncached/1e6*pi + cached/1e6*pc + out/1e6*po.
    h_row = 6
    headers = [
        "cell_id", "est_total_uncached_input", "est_total_cached",
        "est_total_output", "cost_usd",
    ]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=h_row, column=col, value=h).font = BOLD

    first = h_row + 1  # row 7
    for i in range(len(CELLS)):
        r = first + i
        ext_r = ext_first + i
        ws.cell(row=r, column=1, value=f"=extrapolation!A{ext_r}")
        ws.cell(row=r, column=2, value=f"=extrapolation!I{ext_r}")  # uncached
        ws.cell(row=r, column=3, value=f"=extrapolation!G{ext_r}")  # cached
        ws.cell(row=r, column=4, value=f"=extrapolation!H{ext_r}")  # output
        ws.cell(
            row=r, column=5,
            value=(f"=B{r}/1000000*$B$2+C{r}/1000000*$B$3+D{r}/1000000*$B$4"),
        )

    last = first + len(CELLS) - 1
    total = last + 1
    ws.cell(row=total, column=1, value="TOTAL").font = BOLD
    for col in (2, 3, 4, 5):
        L = get_column_letter(col)
        ws.cell(
            row=total, column=col, value=f"=SUM({L}{first}:{L}{last})"
        ).font = BOLD

    # Section 3: provider comparison. Each row carries its own prices and
    # computes the grand total cost from the extrapolation TOTAL row.
    p_row = total + 2
    ws.cell(row=p_row, column=1,
            value="Provider comparison (fill prices per 1M tokens)").font = BOLD
    ph_row = p_row + 1
    p_headers = [
        "provider", "price_input", "price_cached_input", "price_output",
        "total_cost_usd",
    ]
    for col, h in enumerate(p_headers, start=1):
        ws.cell(row=ph_row, column=col, value=h).font = BOLD

    uncached = f"extrapolation!$I${ext_total}"
    cached = f"extrapolation!$G${ext_total}"
    output = f"extrapolation!$H${ext_total}"
    providers = ["Provider A", "Provider B", "Provider C"]
    pfirst = ph_row + 1
    for i, name in enumerate(providers):
        r = pfirst + i
        ws.cell(row=r, column=1, value=name)
        # columns B,C,D left blank for the user
        ws.cell(
            row=r, column=5,
            value=(
                f"={uncached}/1000000*B{r}"
                f"+{cached}/1000000*C{r}"
                f"+{output}/1000000*D{r}"
            ),
        )

    ws.freeze_panes = "A2"
    _widths(ws, [40, 26, 22, 18, 18])
    return ws


def build_notes(wb):
    ws = wb.create_sheet("notes")
    ws["A1"] = "Notes"
    ws["A1"].font = BOLD
    lines = [
        "Data source: dev-midi run, 40 questions per cell, MiniMax-M3 via the "
        "Anthropic-compatible endpoint (minimax-anthropic:MiniMax-M3), "
        "stratified by category (10 per category). This supersedes the earlier "
        "10-question dev-mini probe. Values copied from "
        ".agents/research/2026-06-17-minimax-cost-probe.md.",
        "",
        "Caveats:",
        "- 40 questions per cell still carries sampling variance; the "
        "per-question means are not fully stable estimates.",
        "- A single category-4 question is a heavy token outlier and can pull "
        "the agentic means up.",
        "- Only the 4 MiniMax cells were measured. The GLM cells "
        "(glm-4.7-flashx, glm-5.2) are unmeasured.",
        "- The 12-cell full-factorial figure in the report is a flat 3x "
        "placeholder, not a GLM estimate. GLM tokenization and cache behavior "
        "will differ; re-probe GLM before trusting it.",
        "",
        "Cost reasoning:",
        "- Cached input is usually billed at a large discount, so uncached "
        "input plus output drive the real bill. The extrapolation sheet tracks "
        "uncached input separately (est_total_input minus est_total_cached).",
        "- Output tokens are small everywhere (under 4k mean per question) and "
        "are not the cost driver. Input dominates by 30-40x.",
        "- Paradigm is the dominant axis: agentic is roughly 10x the input of "
        "static per question. Tools-vs-none is second-order and only bites "
        "under agentic.",
        "",
        "How to use:",
        "- extrapolation sheet: edit B1 (testset_size) to rescale all "
        "projections.",
        "- pricing_calc sheet: enter per-1M-token prices in the active price "
        "set (B2:B4) for per-cell costs, or in the provider comparison rows for "
        "side-by-side totals. Leave blank to skip.",
    ]
    for i, line in enumerate(lines, start=2):
        c = ws.cell(row=i, column=1, value=line)
        c.alignment = WRAP
    ws.column_dimensions["A"].width = 100
    return ws


def main():
    wb = Workbook()
    build_per_cell(wb)
    _, ext_first, ext_last, ext_total = build_extrapolation(wb)
    build_pricing(wb, ext_first, ext_last, ext_total)
    build_notes(wb)
    wb.save(OUT)
    print(f"WROTE {OUT}")

    # Validate: reopen and confirm sheets, per_cell values, and formulas.
    wb2 = load_workbook(OUT)
    print("SHEETS:", wb2.sheetnames)

    pc = wb2["per_cell"]
    print("PER_CELL ROWS:")
    for row in pc.iter_rows(min_row=2, max_row=5, values_only=True):
        print(" ", row)

    print("FORMULA CHECKS:")
    checks = [
        ("per_cell!H2", pc["H2"].value),
        ("extrapolation!B4", wb2["extrapolation"]["B4"].value),
        ("extrapolation!F4", wb2["extrapolation"]["F4"].value),
        ("extrapolation!I4", wb2["extrapolation"]["I4"].value),
        ("extrapolation!F8(TOTAL)", wb2["extrapolation"]["F8"].value),
        ("pricing_calc!E7", wb2["pricing_calc"]["E7"].value),
        ("pricing_calc!E11(TOTAL)", wb2["pricing_calc"]["E11"].value),
    ]
    for name, val in checks:
        is_formula = isinstance(val, str) and val.startswith("=")
        print(f"  {name}: {val!r} formula={is_formula}")


if __name__ == "__main__":
    main()
